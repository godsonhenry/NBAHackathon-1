import sys
import argparse

import openpyxl
from openpyxl import cell

class Elimination(object):
    """
    Class that evaluates Analytics_Attachment.xsxl and calculates the date NBA teams were eliminated
    from the playoff contention
    """
    def __init__(self, xlsxFile, divisionSheet, scoresSheet):
		self.xlsx = xlsxFile
		self.workbook = openpyxl.load_workbook(self.xlsx)
		self.division_info = self.workbook.get_sheet_by_name(divisionSheet)
		self.scores = self.workbook.get_sheet_by_name(scoresSheet)

    def generateTeamInfo(self):
        """ Iterates through the team and game info and pulls out necessary information"""
        self.eastTeams = []
        self.westTeams = []
        self.numGames = 0
        #make list of teams and their divisions in east and west conferences
        for rownum in range(self.division_info.max_row - 1):
		    name = self.division_info.cell(row=rownum + 2, column=1).value.encode("utf-8")
		    division = self.division_info.cell(row=rownum + 2, column=2).value.encode("utf-8")
		    conference = self.division_info.cell(row=rownum + 2, column=3).value.encode("utf-8")
		    if conference == 'East':
		        self.eastTeams.append({'name': name, 'division': division, 'wins': 0, 'losses': 0, 'games': []})
		    else:
		    	self.westTeams.append({'name': name, 'division': division, 'wins': 0, 'losses': 0, 'games': []})
        self.currentEightPosE = None
        self.currentEightPosW = None
        self.eastTeamsIndexer = dict((t['name'], i) for i, t in enumerate(self.eastTeams))
        self.westTeamsIndexer = dict((t['name'], i) for i, t in enumerate(self.westTeams))
        self.eliminationDates = []
        self.eastEliminated = 0
        self.westEliminated = 0
        #goes through all of the games and updates team info and elimination dates
        for rownum in range(self.scores.max_row - 1):
            self.numGames += 1
            date = self.scores.cell(row=rownum + 2, column=1).value
            home = self.scores.cell(row=rownum + 2, column=2).value.encode("utf-8")
            away = self.scores.cell(row=rownum + 2, column=3).value.encode("utf-8")
            winner = self.scores.cell(row=rownum + 2, column=6).value.encode("utf-8")
            game = {'home': home, 'away': away, 'winner': winner}
            self.updateRecord(home, "Home", winner, game)
            self.updateRecord(away, "Away", winner, game)
            self.rankEast()
            self.rankWest()
            if len(self.eliminationDates) < 14:
                self.checkElimination('e', date)
                self.checkElimination('w', date)
    def updateRecord(self, team, location, winner, game):
        """Updates the record of a team"""
        ind = self.eastTeamsIndexer.get(team, -1)
        if ind != -1:
            self.eastTeams[ind]['games'].append(game)
            if location == winner:
                self.eastTeams[ind]['wins'] += 1
            else:
                self.eastTeams[ind]['losses'] += 1
        else:
            ind = self.westTeamsIndexer.get(team)
            self.westTeams[ind]['games'].append(game)
            if location == winner:
                self.westTeams[ind]['wins'] += 1
            else:
                self.westTeams[ind]['losses'] += 1
    def checkElimination(self, conf, date):
        """Checks if there is a team to be eliminated in the given conference"""
        teams = []
        eight = None
        if conf == "w" and self.westEliminated < 7:
            teams = self.westTeams
            eight = self.currentEightPosW
        elif conf == "e" and self.eastEliminated < 7:
            teams = self.eastTeams
            eight = self.currentEightPosE
        else:
            return
        for team in teams:
            diffWins = eight['wins'] - team['wins']
            gamesLeft = 82 - team['wins'] - team['losses']
            if any(team['name'] in t for t in self.eliminationDates):
                continue
            #this means there is no way the team can catch up to the eighth ranked team in their conference
            if diffWins > gamesLeft:
                self.eliminationDates.append((team['name'], date))
                if conf == "w":
                    self.westEliminated += 1
                else:
                    self.eastEliminated += 1
            elif diffWins == gamesLeft:
                #check if team could beat eight ranked in tiebreaker
                tiebreakerLoser = self.breakTiebreaker(conf, team, eight, date)
                if tiebreakerLoser[0] == team['name']:
                    self.eliminationDates.append((team['name'], date))
                    if conf == "w":
                        self.westEliminated += 1
                    else:
                        self.eastEliminated += 1
    def rankEast(self):
        """Ranks the teams in the east conference from best to worst"""
        self.eastTeams = sorted(self.eastTeams, key=lambda k: k['losses'])
        self.currentEightPosE = self.eastTeams[7]
        self.eastTeamsIndexer = dict((t['name'], i) for i, t in enumerate(self.eastTeams))
    def rankWest(self):
        """Ranks the teams in the west conference from best to worst"""
        self.westTeams = sorted(self.westTeams, key=lambda k: k['losses'])
        self.currentEightPosW = self.westTeams[7]
        self.westTeamsIndexer = dict((t['name'], i) for i, t in enumerate(self.westTeams))
    def breakTiebreaker(self, conf, first, second, date):
        """Breaks the tiebreaker in the given conference"""
        teams = []
        firstInd = -1
        secondInd = -1
        if conf == 'w':
            teams = self.westTeams
            firstInd = self.westTeamsIndexer.get(first['name'], -1)
            secondInd = self.westTeamsIndexer.get(second['name'], -1)
        else:
            teams = self.eastTeams
            firstInd = self.eastTeamsIndexer.get(first['name'], -1)
            secondInd = self.eastTeamsIndexer.get(second['name'], -1)
        #a tie breaker only needs to be broken if the teams are fighting for the eight spot
        totalGames = 0
        firstTeamWin = 0
        #calculates the teams record on only games they played against eachother
        for game in first['games']:
            if game['home'] == second['name']:
                totalGames += 1
                if game['winner'] == "Away":
                    firstTeamWin += 1
            if game['away'] == second['name']:
                totalGames += 1
                if game['winner'] == "Home":
                    firstTeamWin += 1
        if firstTeamWin < (totalGames - firstTeamWin):
            return self.eliminate(first, conf, firstInd, date)
        elif firstTeamWin > (totalGames - firstTeamWin):
            return self.eliminate(second, conf, secondInd, date)
        #finds if division leader
        firstDiv = first['division']
        secondDiv = second['division']
        firstDivTeams = []
        secondDivTeams = []
        for team in teams:
            if team['division'] == firstDiv:
                firstDivTeams.append(team)
            if team['division'] == secondDiv:
                secondDivTeams.append(team)
        firstDivTeams = sorted(firstDivTeams, key=lambda k: k['losses'])
        secondDivTeams = sorted(secondDivTeams, key=lambda k: k['losses'])
        if firstDivTeams[0] == first and secondDivTeams[0] != second:
            return self.eliminate(second, conf, secondInd, date)
        else: #secondDivTeams[0] == second and firstDivTeams[0] != first:
            return self.eliminate(first, conf, firstInd, date)
    def eliminate(self, team, conf, ind, date):
        """Eliminates a team that was a tiebreaker"""
        year = str(date.year)
        month = str(date.month)
        day = str(date.day)
        writeDate = month + '/' + day + '/' + year
        return (team['name'], writeDate, ind)
    def writeEliminationDates(self):
        """Writes the elimination dates to the excel file"""
        self.generateTeamInfo()
        try:
            self.eliminated_sheet = self.workbook.get_sheet_by_name('Elimination Dates')
            self.workbook.remove_sheet(self.eliminated_sheet)
        except:
            pass
        self.eliminated_sheet = self.workbook.create_sheet('Elimination Dates')
        self.eliminated_sheet['A1'] = 'Team'
        self.eliminated_sheet['B1'] = 'Date Eliminated'
        self.row = 2
        #if there is a tiebreaker, both will be written to eliminationDates, but
        #elimination dates should be only 14 teams
        while len(self.eliminationDates) > 14:
            del self.eliminationDates[len(self.eliminationDates) - 1]
            del self.eliminationDates[len(self.eliminationDates) - 1]
        #writes the eliminated teams
        for team in self.eliminationDates:
            self.eliminated_sheet.cell(row=self.row, column=1).value = team[0]
            year = str(team[1].year)
            month = str(team[1].month)
            day = str(team[1].day)
            date = month + '/' + day + '/' + year
            self.eliminated_sheet.cell(row=self.row, column=2).set_explicit_value(value=date, data_type=cell.Cell.TYPE_STRING)
            self.row += 1
            ind = self.eastTeamsIndexer.get(team[0], -1)
            if ind != -1:
                del self.eastTeams[ind]
                self.eastTeamsIndexer = dict((t['name'], i) for i, t in enumerate(self.eastTeams))
            else:
                ind = self.westTeamsIndexer.get(team[0], -1)
                del self.westTeams[ind]
                self.westTeamsIndexer = dict((t['name'], i) for i, t in enumerate(self.westTeams))
        #ensures there is no tiebreaker left
        if len(self.eastTeams) > 8:
            self.rankEast()
            first = self.eastTeams[7]
            second = self.eastTeams[8]
            date = self.scores.cell(row=self.scores.max_row, column=1).value
            toElim = self.breakTiebreaker('e', first, second, date)
            self.eliminated_sheet.cell(row=self.row, column=1).value = toElim[0]
            self.eliminated_sheet.cell(row=self.row, column=2).set_explicit_value(value=toElim[1], data_type=cell.Cell.TYPE_STRING)
            self.row += 1
            del self.eastTeams[toElim[2]]
        if len(self.westTeams) > 8:
            self.rankWest()
            first = self.westTeams[7]
            second = self.westTeams[8]
            date = self.scores.cell(row=self.scores.max_row, column=1).value
            toElim = self.breakTiebreaker('w', first, second, date)
            self.eliminated_sheet.cell(row=self.row, column=1).value = toElim[0]
            self.eliminated_sheet.cell(row=self.row, column=2).set_explicit_value(value=toElim[1], data_type=cell.Cell.TYPE_STRING)
            self.row += 1
            del self.eastTeams[toElim[2]]
        #writes the remaining teams as Playoffs
        for team in self.eastTeams:
            self.eliminated_sheet.cell(row=self.row, column=1).value = team['name']
            self.eliminated_sheet.cell(row=self.row, column=2).value = "Playoffs"
            self.row += 1
        for team in self.westTeams:
            self.eliminated_sheet.cell(row=self.row, column=1).value = team['name']
            self.eliminated_sheet.cell(row=self.row, column=2).value = "Playoffs"
            self.row += 1
        self.workbook.save(self.xlsx)
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='''Tool to calculate the dates an NBA team was eliminated from the playoff contention'''
    )
    parser.add_argument('--excel_file', type=file, default='Analytics_Attachment.xlsx',
        help='''Excel file with team names, conferences, and all of the scores''')
    parser.add_argument('division_sheet', default='Division_Info',
        help='''Sheet name with at least all of the team names and conference names''')
    parser.add_argument('scores_sheet', default='2016_17_NBA_Scores',
        help='''Sheet name with all of the games from the playoff contention''')
    args = parser.parse_args()
    eliminator = Elimination(args.excel_file.name, args.division_sheet, args.scores_sheet)
    eliminator.writeEliminationDates()
